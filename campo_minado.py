import tkinter as tk
from tkinter import ttk, messagebox
import random
from openpyxl import Workbook


def criar_tabuleiro(linhas, colunas):
    """Cria um tabuleiro vazio com as dimensões especificadas"""
    return [[0 for _ in range(colunas)] for _ in range(linhas)]


def posicionar_minhas(tabuleiro, linhas, colunas):
    """Posiciona minas aleatoriamente no tabuleiro"""
    total_celulas = linhas * colunas
    num_minhas = int(total_celulas * 0.15)  # 15% das células serão minas
    minas_colocadas = 0

    while minas_colocadas < num_minhas:
        linha = random.randint(0, linhas - 1)
        coluna = random.randint(0, colunas - 1)
        if tabuleiro[linha][coluna] != '*':
            tabuleiro[linha][coluna] = '*'
            minas_colocadas += 1

    return tabuleiro


def contar_minhas_adjacentes(tabuleiro, linha, coluna, linhas, colunas):
    """Conta o número de minas nas células adjacentes"""
    contador = 0
    for i in range(max(0, linha - 1), min(linhas, linha + 2)):
        for j in range(max(0, coluna - 1), min(colunas, coluna + 2)):
            if tabuleiro[i][j] == '*':
                contador += 1
    return contador


def calcular_numeros(tabuleiro, linhas, colunas):
    """Calcula os números para as células sem minas"""
    tabuleiro_numerico = criar_tabuleiro(linhas, colunas)

    for i in range(linhas):
        for j in range(colunas):
            if tabuleiro[i][j] == '*':
                tabuleiro_numerico[i][j] = '*'
            else:
                tabuleiro_numerico[i][j] = contar_minhas_adjacentes(tabuleiro, i, j, linhas, colunas)

    return tabuleiro_numerico


def salvar_em_excel(tabuleiro_oculto, tabuleiro_visivel, nome_arquivo):
    """Salva ambos os tabuleiros em um arquivo Excel"""
    wb = Workbook()

    # Criar aba oculta
    aba_oculta = wb.active
    aba_oculta.title = "Oculta"
    for i, linha in enumerate(tabuleiro_oculto, 1):
        for j, celula in enumerate(linha, 1):
            aba_oculta.cell(row=i, column=j, value=str(celula))

    # Criar aba visível
    aba_visível = wb.create_sheet("Visível")
    for i, linha in enumerate(tabuleiro_visivel, 1):
        for j, celula in enumerate(linha, 1):
            aba_visível.cell(row=i, column=j, value=str(celula))

    wb.save(nome_arquivo)


def revelar_celula(event, i, j, botao, tabuleiro_oculto, tabuleiro_visivel, linhas, colunas, root):
    """Revela uma célula quando clicada"""
    if tabuleiro_oculto[i][j] == '*':
        botao.config(text='💣', bg='red', fg='white')
        salvar_em_excel(tabuleiro_oculto, tabuleiro_visivel, "campo_minado.xlsx")
        resposta = messagebox.askyesno("Fim de Jogo", "Você clicou em uma mina! Fim de Jogo! Deseja reiniciar?")
        if resposta:
            root.destroy()
            iniciar_interface()
        else:
            root.destroy()
    else:
        valor = tabuleiro_oculto[i][j]
        botao.config(text=str(valor) if valor > 0 else '', state='disabled', bg='lightgray')
        tabuleiro_visivel[i][j] = valor
        salvar_em_excel(tabuleiro_oculto, tabuleiro_visivel, "campo_minado.xlsx")


def iniciar_jogo(linhas, colunas, root):
    """Inicializa a interface gráfica do jogo"""
    tabuleiro = criar_tabuleiro(linhas, colunas)
    tabuleiro = posicionar_minhas(tabuleiro, linhas, colunas)
    tabuleiro_oculto = calcular_numeros(tabuleiro, linhas, colunas)
    tabuleiro_visivel = [['X' for _ in range(colunas)] for _ in range(linhas)]

    # Criar nova janela para o jogo
    root.destroy()
    jogo_root = tk.Tk()
    jogo_root.title("Campo Minado")
    jogo_root.configure(bg="#2c3e50")

    # Adicionar título
    titulo = tk.Label(jogo_root, text="Campo Minado", font=("Arial", 18, "bold"), fg="white", bg="#2c3e50")
    titulo.pack(pady=10)

    # Frame para o tabuleiro
    frame_tabuleiro = tk.Frame(jogo_root, bg="#34495e", padx=5, pady=5)
    frame_tabuleiro.pack()

    # Criar grid de botões
    for i in range(linhas):
        for j in range(colunas):
            botao = tk.Button(
                frame_tabuleiro, text='', width=4, height=2,
                font=("Arial", 10, "bold"), bg="#bdc3c7", fg="#2c3e50"
            )
            botao.grid(row=i, column=j, padx=1, pady=1)
            botao.bind('<Button-1>', lambda e, x=i, y=j, b=botao: revelar_celula(e, x, y, b, tabuleiro_oculto, tabuleiro_visivel, linhas, colunas, jogo_root))

    jogo_root.mainloop()


def iniciar_interface():
    """Inicializa a interface gráfica para pedir dimensões do tabuleiro"""
    root = tk.Tk()
    root.title("Configuração do Jogo")
    root.configure(bg="#2c3e50")

    # Título
    titulo = tk.Label(root, text="Bem-vindo ao Campo Minado!", font=("Arial", 18, "bold"), fg="white", bg="#2c3e50")
    titulo.grid(row=0, column=0, columnspan=2, pady=10)

    # Inputs de configuração
    tk.Label(root, text="Número de Linhas:", font=("Arial", 12), fg="white", bg="#2c3e50").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    tk.Label(root, text="Número de Colunas:", font=("Arial", 12), fg="white", bg="#2c3e50").grid(row=2, column=0, padx=10, pady=5, sticky="e")

    entrada_linhas = ttk.Entry(root, width=10)
    entrada_linhas.grid(row=1, column=1, padx=10, pady=5)
    entrada_colunas = ttk.Entry(root, width=10)
    entrada_colunas.grid(row=2, column=1, padx=10, pady=5)

    # Botão para iniciar o jogo
    def iniciar():
        try:
            linhas = int(entrada_linhas.get())
            colunas = int(entrada_colunas.get())
            if linhas <= 0 or colunas <= 0:
                raise ValueError
            iniciar_jogo(linhas, colunas, root)
        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira números inteiros positivos.")

    iniciar_btn = ttk.Button(root, text="Iniciar Jogo", command=iniciar)
    iniciar_btn.grid(row=3, column=0, columnspan=2, pady=10)

    root.mainloop()


if __name__ == "__main__":
    iniciar_interface()
